import os
import shutil
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill

# ✅ Cấu hình
excel_path = r"D:\Dữ liệu về tiếng ho NCKH\COUGHVID Dataset\Data đã lọc và sắp xếp 1.xlsx"
sheet_name = "lower_infection"
columns_to_read = ["C", "AA", "AY", "BW"]
source_folder = r"D:\Dữ liệu về tiếng ho NCKH\COUGHVID Dataset\coughvid_20211012"
destination_folder = r"D:\Dữ liệu về tiếng ho NCKH\COUGHVID Dataset\Dữ liệu Coughvid đã lọc\lower_infection1"
output_excel_path = excel_path  # 📌 Ghi đè file gốc

# ✅ Tạo thư mục đích nếu chưa có
os.makedirs(destination_folder, exist_ok=True)

# ✅ Mở file Excel
wb = openpyxl.load_workbook(excel_path)
ws = wb[sheet_name]

# ✅ Chuẩn bị các cột UUID và danh sách đã sao chép
uuid_success = set()
uuid_to_cells = {}

for col_letter in columns_to_read:
    col_idx = column_index_from_string(col_letter)
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if cell.value:
            uuid = str(cell.value).strip()
            uuid_to_cells.setdefault(uuid, []).append(cell)

# ✅ Sao chép file
extensions = [".webm", ".json", ".wav"]
copied, missing = 0, 0

for uuid in uuid_to_cells:
    copied_any = False
    for ext in extensions:
        filename = f"{uuid}{ext}"
        src = os.path.join(source_folder, filename)
        dst = os.path.join(destination_folder, filename)
        if os.path.exists(src):
            shutil.copy2(src, dst)
            print(f"✅ Đã sao chép: {filename}")
            copied += 1
            copied_any = True
        else:
            print(f"❌ Không tìm thấy: {filename}")
            missing += 1
    if copied_any:
        uuid_success.add(uuid)

# ✅ Tô màu ô thành công
fill_red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
colored_cells = 0

for uuid in uuid_success:
    for cell in uuid_to_cells.get(uuid, []):
        cell.fill = fill_red
        colored_cells += 1

# ✅ Ghi đè file Excel gốc
wb.save(output_excel_path)

# ✅ Tổng kết
print(f"\n🎉 Đã sao chép {copied} file, {missing} file không tìm thấy.")
print(f"📘 Đã tô màu {colored_cells} ô UUID trong file Excel.")
print(f"💾 File Excel đã được cập nhật và ghi đè.")
